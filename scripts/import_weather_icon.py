#!/usr/bin/env python3
"""
导入天气现象映射表
将 downloader/weather_icon.xlsx 导入到 weather.weather_icon 表

用法：
  python scripts/import_weather_icon.py
"""

import sys
from pathlib import Path

import openpyxl
import pymysql

# 项目根目录
ROOT = Path(__file__).resolve().parent.parent

# 数据库配置（与 .env.example 一致）
DB_CONFIG = {
    "host": "127.0.0.1",
    "port": 3307,
    "user": "root",
    "password": "123456",
    "database": "weather",
    "charset": "utf8mb4",
}


def read_xlsx(path: Path) -> list[dict]:
    """读取 xlsx 文件"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # 第一行是表头
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        rows.append(record)

    wb.close()
    return rows


def import_to_db(rows: list[dict]):
    """导入到数据库"""
    conn = pymysql.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:
            # 清空旧数据
            cur.execute("TRUNCATE TABLE weather_icon")
            print("✅ 已清空 weather_icon 表")

            # 批量插入
            sql = """
                INSERT INTO weather_icon
                (id, icon_day, icon_night, condition_zh, condition_en,
                 condition_tw, condition_hk, condition_es, condition_pt,
                 condition_ru, condition_r, condition_de, condition_ar,
                 condition_ko, condition_ja, condition_hi, condition_te, weather_id)
                VALUES
                (%(id)s, %(icon_day)s, %(icon_night)s, %(condition_zh)s, %(condition_en)s,
                 %(condition_tw)s, %(condition_hk)s, %(condition_es)s, %(condition_pt)s,
                 %(condition_ru)s, %(condition_r)s, %(condition_de)s, %(condition_ar)s,
                 %(condition_ko)s, %(condition_ja)s, %(condition_hi)s, %(condition_te)s, %(weather_id)s)
            """
            cur.executemany(sql, rows)
            conn.commit()
            print(f"✅ 导入完成，共 {len(rows)} 条记录")

    finally:
        conn.close()


def main():
    xlsx_path = ROOT / "downloader" / "weather_icon.xlsx"
    if not xlsx_path.exists():
        print(f"❌ 文件不存在: {xlsx_path}")
        sys.exit(1)

    print(f"📖 读取 {xlsx_path} ...")
    rows = read_xlsx(xlsx_path)
    print(f"   共 {len(rows)} 条数据")

    print(f"💾 导入到 {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']} ...")
    import_to_db(rows)


if __name__ == "__main__":
    main()
